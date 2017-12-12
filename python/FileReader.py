import openpyxl
import pprint

path = r'C:\Users\willi\Documents\test.xlsx'
# open workbook with only data in cells
wb = openpyxl.load_workbook(path, data_only=True)
# get first worksheet name
wsn = wb.get_sheet_names()[0]
# so I can get worksheet
ws = wb.get_sheet_by_name(wsn)

employeeList = {}
# loop thru row and columns
for row in ws.iter_rows('A{}:J{}'.format(ws.min_row+1,ws.max_row)):
    employee = row[0].value.upper()
    customer = row[2].value
    product = row[4].value
    revenue = row[9].value or 0
    # check if contain keywords
    if any(p in product for p in ('硬化劑', '發泡劑', '設備', '薄膜')):
        #check if employee already added
        if not any(emp for emp in employeeList if emp == employee):
            employeeList[employee] = [{customer:{product:revenue}}]
        else:
            #check if customer already added
            salesReport = employeeList[employee]
            #if there is no existing customer
            if not any(cus for cus in salesReport if customer in cus):
                salesReport.append({customer:{product:revenue}})
            else:
                cus = next(cus for cus in salesReport if customer in cus)
                #if there is no existing product
                if not any(prod for prod in cus.get(customer) if product == prod):
                    cus[customer][product] = revenue
                else:
                    cus[customer][product] += revenue
pprint.pprint(employeeList)
