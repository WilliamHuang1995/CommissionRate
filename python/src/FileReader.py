import openpyxl

def processData(ws):
    employeeList = {}
    # loop thru row and columns
    for row in ws.iter_rows('A{}:J{}'.format(ws.min_row+1,ws.max_row)):
        employee = row[0].value.title()
        customer = row[2].value
        product = row[4].value
        revenue = row[9].value or 0
        # check if contain keywords
        if any(p in product for p in ('硬化劑', '發泡劑', '設備', '薄膜')):
            #check if employee already added
            if not any(emp for emp in employeeList if emp == employee):
                employeeList[employee] = [{customer:{product:revenue}}]
            else:
                #check if customer already added
                salesReport = employeeList[employee]
                #if there is no existing customer
                if not any(cus for cus in salesReport if customer in cus):
                    salesReport.append({customer:{product:revenue}})
                else:
                    cus = next(cus for cus in salesReport if customer in cus)
                    #if there is no existing product
                    if not any(prod for prod in cus.get(customer) if product == prod):
                        cus[customer][product] = revenue
                    else:
                        cus[customer][product] += revenue
    return employeeList

def outputData(dictionary):
    ## for each employee listed
    for employee in dictionary:
        ows = owb.create_sheet(title=employee)
        #insert header
        for col in range(1,len(header)):
            ows.cell(row=1,column=col).value = header[col]
        ows.cell(row=2,column=1).value = employee
        #row starts at 2
        #column starts at 2
        row = 2
        
        #iterate thru companies
        for companies in dictionary[employee]:
            col = 3
            for company in companies:
                ows.cell(row=row,column=2).value = company
                for product in companies[company]:
                    revenue = companies[company][product]
                    values = [product,revenue,revenue*0.02,revenue*0.03,revenue*0.05]
                    for i in range(len(values)):
                        ows.cell(row=row,column=col+i).value = values[i]
                    row += 1

path = r'C:\Users\willi\Documents\test.xlsx'
dest = r'C:\Users\willi\Documents\PythonOutput.xlsx'
header = ['','業務','客戶','產品','金額','2%佣金','3%佣金','5%佣金','Subtotal']
# open workbook with only data in cells
wb = openpyxl.load_workbook(path, data_only=True)
# get first worksheet name
wsn = wb.get_sheet_names()[0]
# so I can get worksheet
ws = wb.get_sheet_by_name(wsn)
dictionary = processData(ws)
## create workbook
owb = openpyxl.Workbook()
outputData(dictionary)
#remove default sheet that Workbook creates
owb.remove_sheet(owb.get_sheet_by_name('Sheet'))
owb.save(filename=dest)
