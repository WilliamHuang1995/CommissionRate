import openpyxl
from tkinter import *
from tkinter import filedialog


class GUI:
    def __init__(self, master):
        self.master = master
        master.title("Commission Calculator")

        self.select_button = Button(master, text="Select Input", command=self.select_input).grid(row=0)
        self.input_label = Label(master, text="", anchor="w", width=50)
        self.input_label.grid(row=0, column=1)
        self.output_button = Button(master, text="Select Output", command=self.select_output).grid(row=1)
        self.output_label = Label(master, text="", anchor="w", width=50)
        self.output_label.grid(row=1, column=1)
        self.run_button = Button(master, text="Run!", command=main).grid(row=2)
        self.close_button = Button(master, text="Close", command=master.quit).grid(row=2, column=1)

    def select_input(self):
        global path
        fn = filedialog.askopenfilename(initialdir="/", title="Select file",
                                        filetypes=(("Excel 2010 Files", "*.xlsx"), ("All Files", "*.*")))
        self.input_label['text'] = fn
        path = self.input_label['text']
        return fn

    def select_output(self):
        global dest
        fn = filedialog.askdirectory()
        self.output_label['text'] = fn + '/Output.xlsx'
        dest = self.output_label['text']
        return fn


def main():
    global my_gui
    if not path or not dest:
        print('input or output not set properly!')
        return
    # open workbook with only data in cells
    wb = openpyxl.load_workbook(path, data_only=True)
    # get first worksheet name
    wsn = wb.get_sheet_names()[0]
    # so I can get worksheet
    ws = wb.get_sheet_by_name(wsn)
    dictionary = process_data(ws)
    # create workbook
    owb = openpyxl.Workbook()
    output_data(dictionary, owb)
    # remove default sheet that Workbook creates
    owb.remove_sheet(owb.get_sheet_by_name('Sheet'))
    owb.save(filename=dest)
    my_gui.output_label['text'] = 'Success'
    my_gui.input_label['text'] = 'Success'


def process_data(ws):
    employee_dict = {}
    # loop through row and columns
    for row in ws.iter_rows('A{}:J{}'.format(ws.min_row + 1, ws.max_row)):
        employee = row[0].value.title()
        customer = row[2].value
        product = row[4].value
        revenue = row[9].value or 0
        # check if contain keywords
        if any(p in product for p in ('硬化劑', '發泡劑', '設備', '薄膜')):
            # check if employee already added
            # if not any(emp for emp in employee_dict if emp == employee):
            if employee not in employee_dict.keys():
                employee_dict[employee] = [{customer: {product: revenue}}]
            else:
                # check if customer already added
                sales_report = employee_dict[employee]
                # if there is no existing customer
                if not any(cus for cus in sales_report if customer in cus):
                    sales_report.append({customer: {product: revenue}})
                else:
                    cus = next(cus for cus in sales_report if customer in cus)
                    # if there is no existing product
                    if not any(prod for prod in cus.get(customer) if product == prod):
                        cus[customer][product] = revenue
                    else:
                        cus[customer][product] += revenue
    return employee_dict


def output_data(dictionary, owb):
    header = ['', '業務', '客戶', '產品', '金額', '2%佣金', '3%佣金', '5%佣金', 'Subtotal']
    # for each employee listed
    for employee in dictionary:
        ows = owb.create_sheet(title=employee)
        # insert header
        for col in range(1, len(header)):
            ows.cell(row=1, column=col).value = header[col]
        ows.cell(row=2, column=1).value = employee
        # row starts at 2
        # column starts at 2
        row = 2

        # iterate thru companies
        for companies in dictionary[employee]:
            col = 3
            for company in companies:
                ows.cell(row=row, column=2).value = company
                for product in companies[company]:
                    revenue = companies[company][product]
                    values = [product, revenue, revenue * 0.02, revenue * 0.03, revenue * 0.05]
                    for i in range(len(values)):
                        ows.cell(row=row, column=col + i).value = values[i]
                    row += 1


path = ''
dest = ''
root = Tk()
my_gui = GUI(root)
root.mainloop()
root.destroy()
